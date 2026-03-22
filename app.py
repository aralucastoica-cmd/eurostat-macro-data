"""
Eurostat Macroeconomic Data Downloader
Descarca date macroeconomice de la Eurostat cu un singur click → Excel formatat
"""

from flask import Flask, render_template, request, send_file, jsonify
import requests
import pandas as pd
import io
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

app = Flask(__name__)

BASE_URL = "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data"

GEO_LABELS = {
    # ── Agregate ──────────────────────────────────────────────────────────────
    "EU27_2020": "UE-27 (medie)",
    "EA20":      "Zona Euro (20)",
    # ── Tari membre UE (27) ───────────────────────────────────────────────────
    "AT": "Austria",
    "BE": "Belgia",
    "BG": "Bulgaria",
    "CY": "Cipru",
    "CZ": "Cehia",
    "DE": "Germania",
    "DK": "Danemarca",
    "EE": "Estonia",
    "ES": "Spania",
    "FI": "Finlanda",
    "FR": "Franta",
    "GR": "Grecia",
    "HR": "Croatia",
    "HU": "Ungaria",
    "IE": "Irlanda",
    "IT": "Italia",
    "LT": "Lituania",
    "LU": "Luxemburg",
    "LV": "Letonia",
    "MT": "Malta",
    "NL": "Olanda",
    "PL": "Polonia",
    "PT": "Portugalia",
    "RO": "Romania",
    "SE": "Suedia",
    "SI": "Slovenia",
    "SK": "Slovacia",
    # ── Tari candidate UE (date disponibile Eurostat) ─────────────────────────
    "AL": "Albania",
    "BA": "Bosnia-Hertegovina",
    "GE": "Georgia",
    "MD": "Moldova",
    "ME": "Muntenegru",
    "MK": "Macedonia de Nord",
    "RS": "Serbia",
    "TR": "Turcia",
    "UA": "Ucraina",
    "XK": "Kosovo",
    # ── EEA / EFTA ────────────────────────────────────────────────────────────
    "CH": "Elvetia",
    "IS": "Islanda",
    "LI": "Liechtenstein",
    "NO": "Norvegia",
    # ── Alte tari europene ────────────────────────────────────────────────────
    "GB": "Marea Britanie (UK)",
    "RS": "Serbia",
}

# ─────────────────────────────────────────────────────────────────────────────
# LISTA INDICATORI  (dataset + filtre Eurostat API)
# ─────────────────────────────────────────────────────────────────────────────
INDICATORS = [
    # ── PIB & CRESTERE ────────────────────────────────────────────────────────
    {"category": "PIB & Crestere Economica",
     "id": "gdp_meur",     "name": "PIB nominal (mil. EUR)",
     "dataset": "nama_10_gdp",
     "filters": {"unit": "CP_MEUR",      "na_item": "B1GQ", "freq": "A"}},

    {"category": "PIB & Crestere Economica",
     "id": "gdp_growth",   "name": "Crestere PIB real (%)",
     "dataset": "nama_10_gdp",
     "filters": {"unit": "CLV_PCH_PRE",  "na_item": "B1GQ", "freq": "A"}},

    {"category": "PIB & Crestere Economica",
     "id": "gdp_per_cap",  "name": "PIB per capita (EUR)",
     "dataset": "nama_10_pc",
     "filters": {"unit": "CP_EUR_HAB",   "na_item": "B1GQ", "freq": "A"}},

    {"category": "PIB & Crestere Economica",
     "id": "gdp_pc_growth","name": "PIB per capita raportat la UE27 (%)",
     "dataset": "nama_10_pc",
     "filters": {"unit": "PC_EU27_2020_HAB_MEUR_CP", "na_item": "B1GQ", "freq": "A"}},

    {"category": "PIB & Crestere Economica",
     "id": "gdp_pps",      "name": "PIB per capita PPS (EU27=100)",
     "dataset": "sdg_10_10",
     "filters": {"indic_ppp": "VI_PPS_EU27_2020_HAB", "ppp_cat18": "GDP", "unit": "PC"}},

    # ── INFLATIE & PRETURI ─────────────────────────────────────────────────────
    {"category": "Inflatie & Preturi",
     "id": "hicp_rate",    "name": "Inflatie HICP – rata anuala (%)",
     "dataset": "prc_hicp_aind",
     "filters": {"unit": "RCH_A_AVG",    "coicop": "CP00"}},

    {"category": "Inflatie & Preturi",
     "id": "hicp_idx",     "name": "Indice HICP (2015=100)",
     "dataset": "prc_hicp_aind",
     "filters": {"unit": "INX_A_AVG",    "coicop": "CP00"}},

    {"category": "Inflatie & Preturi",
     "id": "hicp_core",    "name": "Inflatie de baza excl. energie & alimente (%)",
     "dataset": "prc_hicp_aind",
     "filters": {"unit": "RCH_A_AVG",    "coicop": "TOT_X_NRG_FOOD"}},

    {"category": "Inflatie & Preturi",
     "id": "hicp_energy",  "name": "Inflatie energie (%)",
     "dataset": "prc_hicp_aind",
     "filters": {"unit": "RCH_A_AVG",    "coicop": "CP04"}},

    {"category": "Inflatie & Preturi",
     "id": "hicp_food",    "name": "Inflatie alimente & bauturi nealcoolice (%)",
     "dataset": "prc_hicp_aind",
     "filters": {"unit": "RCH_A_AVG",    "coicop": "CP011"}},

    # ── PIATA MUNCII ─────────────────────────────────────────────────────────
    {"category": "Piata Muncii",
     "id": "unemployment", "name": "Rata somajului total (%)",
     "dataset": "une_rt_a",
     "filters": {"unit": "PC_ACT", "sex": "T", "age": "Y15-74"}},

    {"category": "Piata Muncii",
     "id": "unemp_youth",  "name": "Somaj tineri 15-24 ani (%)",
     "dataset": "une_rt_a",
     "filters": {"unit": "PC_ACT", "sex": "T", "age": "Y15-24"}},

    {"category": "Piata Muncii",
     "id": "unemp_female", "name": "Somaj femei (%)",
     "dataset": "une_rt_a",
     "filters": {"unit": "PC_ACT", "sex": "F", "age": "Y15-74"}},

    {"category": "Piata Muncii",
     "id": "employment",   "name": "Rata de ocupare 20-64 ani (%)",
     "dataset": "lfsi_emp_a",
     "filters": {"unit": "PC_POP", "sex": "T", "age": "Y20-64"}},

    {"category": "Piata Muncii",
     "id": "emp_female",   "name": "Rata de ocupare femei 20-64 ani (%)",
     "dataset": "lfsi_emp_a",
     "filters": {"unit": "PC_POP", "sex": "F", "age": "Y20-64"}},

    # ── IMOBILIARE ───────────────────────────────────────────────────────────
    {"category": "Imobiliare",
     "id": "hpi_total",    "name": "Indice Preturi Locuinte – total (2010=100)",
     "dataset": "prc_hpi_a",
     "filters": {"unit": "I10_A_AVG", "purchase": "TOTAL"}},

    {"category": "Imobiliare",
     "id": "hpi_new",      "name": "Indice Preturi Locuinte – noi (2010=100)",
     "dataset": "prc_hpi_a",
     "filters": {"unit": "I10_A_AVG", "purchase": "DW_NEW"}},

    {"category": "Imobiliare",
     "id": "hpi_existing", "name": "Indice Preturi Locuinte – existente (2010=100)",
     "dataset": "prc_hpi_a",
     "filters": {"unit": "I10_A_AVG", "purchase": "DW_EXST"}},

    {"category": "Imobiliare",
     "id": "hpi_change",   "name": "Variatie HPI an/an (%)",
     "dataset": "prc_hpi_a",
     "filters": {"unit": "RCH_A_AVG", "purchase": "TOTAL"}},

    # ── FINANTE PUBLICE ───────────────────────────────────────────────────────
    {"category": "Finante Publice",
     "id": "gov_deficit",  "name": "Deficit/Surplus guvernamental (% PIB)",
     "dataset": "gov_10dd_edpt1",
     "filters": {"unit": "PC_GDP", "na_item": "B9",  "sector": "S13", "freq": "A"}},

    {"category": "Finante Publice",
     "id": "gov_debt",     "name": "Datorie publica (% PIB)",
     "dataset": "gov_10dd_edpt1",
     "filters": {"unit": "PC_GDP", "na_item": "GD",  "sector": "S13", "freq": "A"}},

    {"category": "Finante Publice",
     "id": "gov_revenue",  "name": "Venituri guvernamentale (% PIB)",
     "dataset": "gov_10a_main",
     "filters": {"unit": "PC_GDP", "na_item": "TR",  "sector": "S13", "freq": "A"}},

    {"category": "Finante Publice",
     "id": "gov_expenditure","name": "Cheltuieli guvernamentale (% PIB)",
     "dataset": "gov_10a_main",
     "filters": {"unit": "PC_GDP", "na_item": "TE",  "sector": "S13", "freq": "A"}},

    # ── DOBANZI & RATE ────────────────────────────────────────────────────────
    {"category": "Dobanzi & Rate",
     "id": "lt_interest",  "name": "Dobanda pe termen lung (benchmark, %)",
     "dataset": "irt_lt_mcby_a",
     "filters": {"int_rt": "MCBY"}},

    # ── BALANTA DE PLATI ─────────────────────────────────────────────────────
    {"category": "Balanta de Plati",
     "id": "curr_acc",     "name": "Cont curent (% PIB)",
     "dataset": "tipsbp20",
     "filters": {"bop_item": "CA", "unit": "PC_GDP", "stk_flow": "BAL",
                 "partner": "WRL_REST", "sector10": "S1", "sectpart": "S1", "s_adj": "NSA"}},

    # ── DEMOGRAFIE ───────────────────────────────────────────────────────────
    {"category": "Demografie",
     "id": "population",   "name": "Populatie totala (persoane)",
     "dataset": "demo_pjan",
     "filters": {"unit": "NR", "sex": "T", "age": "TOTAL"}},

    {"category": "Demografie",
     "id": "pop_growth",   "name": "Spor natural populatie (persoane)",
     "dataset": "demo_gind",
     "filters": {"indic_de": "NATGROW"}},

    {"category": "Demografie",
     "id": "net_migration","name": "Migratie neta (persoane)",
     "dataset": "demo_gind",
     "filters": {"indic_de": "CNMIGRAT"}},

    # ── SOCIAL & INEGALITATE ─────────────────────────────────────────────────
    {"category": "Social & Inegalitate",
     "id": "poverty",      "name": "Rata saraciei dupa transferuri sociale (%)",
     "dataset": "ilc_li02",
     "filters": {"unit": "PC", "sex": "T", "age": "TOTAL", "indic_il": "LI_R_MD60"}},

    {"category": "Social & Inegalitate",
     "id": "gini",         "name": "Coeficient Gini",
     "dataset": "ilc_di12",
     "filters": {"statinfo": "GINI_HND"}},

    {"category": "Social & Inegalitate",
     "id": "arope",        "name": "Risc saracie sau excluziune sociala (%)",
     "dataset": "ilc_peps01n",
     "filters": {"unit": "PC", "sex": "T", "age": "TOTAL"}},

    # ── ENERGIE & MEDIU ───────────────────────────────────────────────────────
    {"category": "Energie & Mediu",
     "id": "renewable",    "name": "Energii regenerabile (% consum final brut)",
     "dataset": "sdg_07_40",
     "filters": {"nrg_bal": "REN", "unit": "PC"}},

    {"category": "Energie & Mediu",
     "id": "ghg_total",    "name": "Emisii CO2 pe cap de locuitor (tone)",
     "dataset": "sdg_13_10",
     "filters": {"src_crf": "TOTXMEMO", "unit": "T_HAB"}},
]

# ─────────────────────────────────────────────────────────────────────────────
# EUROSTAT API CLIENT
# ─────────────────────────────────────────────────────────────────────────────

def parse_eurostat_json(data):
    """Convert Eurostat JSON-stat response → pandas DataFrame."""
    dims    = data.get("id", [])
    sizes   = data.get("size", [])
    dim_info= data.get("dimension", {})
    values  = data.get("value", {})

    if not dims or not values:
        return pd.DataFrame()

    dim_cats = []
    for dim in dims:
        cats     = dim_info[dim]["category"]
        idx_dict = cats.get("index", {})
        lbl_dict = cats.get("label", {})
        if isinstance(idx_dict, list):
            ordered = [(str(x), lbl_dict.get(str(x), str(x))) for x in idx_dict]
        else:
            ordered = [(k, lbl_dict.get(k, k))
                       for k, _ in sorted(idx_dict.items(), key=lambda x: x[1])]
        dim_cats.append(ordered)

    total = 1
    for s in sizes:
        total *= s

    records = []
    for i in range(total):
        val = values.get(str(i)) if str(i) in values else values.get(i)
        if val is None:
            continue

        remaining = i
        indices   = []
        for s in reversed(sizes):
            indices.append(remaining % s)
            remaining //= s
        indices.reverse()

        record = {}
        for j, dim in enumerate(dims):
            if indices[j] < len(dim_cats[j]):
                record[dim] = dim_cats[j][indices[j]][0]
        record["value"] = val
        records.append(record)

    return pd.DataFrame(records)


def fetch_indicator(dataset, filters, geos, year_start, year_end):
    """Fetch one indicator; return (DataFrame | None, error_string | None)."""
    params = {**filters, "format": "JSON", "lang": "en",
              "sinceTimePeriod": str(year_start),
              "untilTimePeriod": str(year_end)}

    # Build query string — geo appears multiple times
    qs_parts = [f"{k}={v}" for k, v in params.items()]
    for geo in geos:
        qs_parts.append(f"geo={geo}")

    url = f"{BASE_URL}/{dataset}?" + "&".join(qs_parts)

    try:
        resp = requests.get(url, timeout=30)
        if resp.status_code != 200:
            return None, f"HTTP {resp.status_code}"
        data = resp.json()
        if "error" in data:
            return None, str(data["error"])
        df = parse_eurostat_json(data)
        return (df, None) if not df.empty else (None, "No data")
    except Exception as e:
        return None, str(e)


def fetch_all_parallel(selected_inds, geos, year_start, year_end):
    """Fetch multiple indicators in parallel (5 threads)."""
    results = [None] * len(selected_inds)

    def task(idx, ind):
        df, err = fetch_indicator(ind["dataset"], ind["filters"], geos, year_start, year_end)
        return idx, df, err

    with ThreadPoolExecutor(max_workers=5) as pool:
        futures = {pool.submit(task, i, ind): i for i, ind in enumerate(selected_inds)}
        for fut in as_completed(futures):
            idx, df, err = fut.result()
            ind = selected_inds[idx]
            results[idx] = {
                "category": ind["category"],
                "name":     ind["name"],
                "df":       df,
                "error":    err,
            }

    return results


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER  (culori brand: navy #0D1B2A + gold #C9A227)
# ─────────────────────────────────────────────────────────────────────────────

NAVY   = "0D1B2A"
GOLD   = "C9A227"
ACCENT = "1A3A5C"
WHITE  = "FFFFFF"
LIGHT  = "F5F7FA"


def style_cell(cell, bg=NAVY, fg=GOLD, size=10, bold=True, align="center"):
    cell.fill      = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font      = Font(color=fg, bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)


def build_excel(results, geos, year_start, year_end):
    wb    = Workbook()
    ws_all = wb.active
    ws_all.title = "Toti Indicatorii"

    thin = Border(
        left   = Side(style="thin", color="DDDDDD"),
        right  = Side(style="thin", color="DDDDDD"),
        top    = Side(style="thin", color="DDDDDD"),
        bottom = Side(style="thin", color="DDDDDD"),
    )

    geo_str  = " | ".join(GEO_LABELS.get(g, g) for g in geos)
    gen_date = datetime.now().strftime("%d.%m.%Y %H:%M")

    # ── Main sheet title ──────────────────────────────────────────────────────
    ws_all.merge_cells("A1:H1")
    style_cell(ws_all["A1"], size=13,
               fg=GOLD, bg=NAVY)
    ws_all["A1"].value = f"Date Macroeconomice Eurostat  |  {geo_str}  |  {year_start}–{year_end}"
    ws_all.row_dimensions[1].height = 32

    ws_all.merge_cells("A2:H2")
    style_cell(ws_all["A2"], bg=GOLD, fg=NAVY, size=9, bold=False)
    ws_all["A2"].value = f"Sursa: Eurostat  |  Generat: {gen_date}"
    ws_all.row_dimensions[2].height = 16

    main_row = 4

    # ── Group results by category ─────────────────────────────────────────────
    by_cat = defaultdict(list)
    for r in results:
        if r and r["df"] is not None and not r["df"].empty:
            by_cat[r["category"]].append(r)

    for category, items in by_cat.items():
        # --- Category sheet ---
        ws_cat = wb.create_sheet(title=category[:31])
        ws_cat.merge_cells("A1:H1")
        style_cell(ws_cat["A1"], size=12)
        ws_cat["A1"].value = category
        ws_cat.row_dimensions[1].height = 28
        cat_row = 3

        # --- Category header on main sheet ---
        ws_all.merge_cells(f"A{main_row}:H{main_row}")
        style_cell(ws_all[f"A{main_row}"], bg=ACCENT, fg=GOLD, size=10, align="left")
        ws_all[f"A{main_row}"].value = f"  {category.upper()}"
        ws_all.row_dimensions[main_row].height = 22
        main_row += 1

        for item in items:
            df = item["df"]
            if "time" not in df.columns or "geo" not in df.columns:
                continue

            try:
                pivot = df.pivot_table(index="time", columns="geo",
                                       values="value", aggfunc="first")
                pivot = pivot.sort_index(ascending=False)
                pivot.columns = [GEO_LABELS.get(c, c) for c in pivot.columns]
                pivot.index.name = "An"
                col_names = list(pivot.columns)

                # ── Write to MAIN sheet ───────────────────────────────────────
                # Indicator label row
                ws_all.merge_cells(f"A{main_row}:H{main_row}")
                c = ws_all[f"A{main_row}"]
                c.value = item["name"]
                c.fill  = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")
                c.font  = Font(bold=True, size=9, color=NAVY)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                ws_all.row_dimensions[main_row].height = 18
                main_row += 1

                # Column headers
                ws_all.cell(row=main_row, column=1, value="An")
                style_cell(ws_all.cell(row=main_row, column=1),
                           bg=ACCENT, fg=WHITE, size=9)
                ws_all.cell(row=main_row, column=1).border = thin
                for ci, cn in enumerate(col_names, 2):
                    style_cell(ws_all.cell(row=main_row, column=ci),
                               bg=ACCENT, fg=WHITE, size=9)
                    ws_all.cell(row=main_row, column=ci).value  = cn
                    ws_all.cell(row=main_row, column=ci).border = thin
                ws_all.row_dimensions[main_row].height = 16
                main_row += 1

                # Data rows
                for ri, (yr, row) in enumerate(pivot.iterrows()):
                    fc = LIGHT if ri % 2 == 0 else WHITE
                    rf = PatternFill(start_color=fc, end_color=fc, fill_type="solid")

                    yc = ws_all.cell(row=main_row, column=1, value=str(yr))
                    yc.fill = rf; yc.font = Font(bold=True, size=9)
                    yc.alignment = Alignment(horizontal="center", vertical="center")
                    yc.border = thin

                    for ci, val in enumerate(row, 2):
                        vc = ws_all.cell(row=main_row, column=ci)
                        if pd.notna(val):
                            vc.value          = round(float(val), 3)
                            vc.number_format  = "#,##0.00"
                        else:
                            vc.value = "—"
                        vc.fill = rf
                        vc.font = Font(size=9)
                        vc.alignment = Alignment(horizontal="right", vertical="center")
                        vc.border = thin
                    ws_all.row_dimensions[main_row].height = 15
                    main_row += 1

                main_row += 1  # blank line

                # ── Write to CATEGORY sheet ───────────────────────────────────
                ws_cat.merge_cells(f"A{cat_row}:H{cat_row}")
                nc = ws_cat[f"A{cat_row}"]
                nc.value = item["name"]
                nc.fill  = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
                nc.font  = Font(color=NAVY, bold=True, size=10)
                nc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                ws_cat.row_dimensions[cat_row].height = 20
                cat_row += 1

                headers = ["An"] + col_names
                for hi, h in enumerate(headers, 1):
                    hc = ws_cat.cell(row=cat_row, column=hi, value=h)
                    style_cell(hc, bg=ACCENT, fg=WHITE, size=9)
                    hc.border = thin
                ws_cat.row_dimensions[cat_row].height = 16
                cat_row += 1

                for ri, (yr, row) in enumerate(pivot.iterrows()):
                    fc = LIGHT if ri % 2 == 0 else WHITE
                    rf = PatternFill(start_color=fc, end_color=fc, fill_type="solid")

                    yc = ws_cat.cell(row=cat_row, column=1, value=str(yr))
                    yc.fill = rf; yc.font = Font(bold=True, size=9)
                    yc.alignment = Alignment(horizontal="center"); yc.border = thin

                    for ci, val in enumerate(row, 2):
                        vc = ws_cat.cell(row=cat_row, column=ci)
                        if pd.notna(val):
                            vc.value         = round(float(val), 3)
                            vc.number_format = "#,##0.00"
                        else:
                            vc.value = "—"
                        vc.fill = rf; vc.font = Font(size=9)
                        vc.alignment = Alignment(horizontal="right"); vc.border = thin
                    ws_cat.row_dimensions[cat_row].height = 15
                    cat_row += 1

                cat_row += 2

            except Exception:
                pass

        # Auto-width category sheet
        for col in ws_cat.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws_cat.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Auto-width main sheet
    for i in range(1, 9):
        ws_all.column_dimensions[get_column_letter(i)].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# FLASK ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    by_cat = defaultdict(list)
    for ind in INDICATORS:
        by_cat[ind["category"]].append({"id": ind["id"], "name": ind["name"]})
    return render_template("index.html", categories=dict(by_cat))


@app.route("/download", methods=["POST"])
def download():
    payload    = request.get_json()
    selected   = set(payload.get("indicators", []))
    geos       = payload.get("geos",       ["RO", "EU27_2020"])
    year_start = int(payload.get("year_start", 2010))
    year_end   = int(payload.get("year_end",   2024))

    chosen = [ind for ind in INDICATORS if ind["id"] in selected] if selected else INDICATORS

    results = fetch_all_parallel(chosen, geos, year_start, year_end)
    buf     = build_excel(results, geos, year_start, year_end)
    fname   = f"Eurostat_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


if __name__ == "__main__":
    import os

    port = int(os.environ.get("PORT", 5000))
    is_cloud = os.environ.get("RENDER") or os.environ.get("RAILWAY_ENVIRONMENT")

    if not is_cloud:
        # Local: deschide browser automat
        import webbrowser, threading, time
        def open_browser():
            time.sleep(1.5)
            webbrowser.open(f"http://localhost:{port}")
        print("\n" + "=" * 52)
        print("  EUROSTAT MACRO DATA DOWNLOADER")
        print(f"  Deschide browser: http://localhost:{port}")
        print("=" * 52 + "\n")
        threading.Thread(target=open_browser, daemon=True).start()

    app.run(host="0.0.0.0", port=port, debug=False)
